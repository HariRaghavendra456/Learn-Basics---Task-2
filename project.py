from openpyxl import Workbook , load_workbook

from openpyxl.utils import get_column_letter

wb1=load_workbook("Input1.xlsx")  #To read an Excel file you have to open the spreadsheet using the load_workbook() method.

wb2 = Workbook() #An empty spreadsheet can be created using the Workbook() method.

ws1= wb1.active #active to select the first sheet available.

ws2= wb2.active
ws2.title = "Output" #To the title for the spreadsheet.

ws2.append(["Name","Username","Chapter Tag","Test_Name","answered","correct","score","skipped","time-taken (seconds)","wrong"])

#Appending the data into the new spreadsheet.


ChapterTag = []

ChapterTest = []

TopicTest  = []

FullChapterTest =  []

column = 1

Column_Names = []

while (ws1[get_column_letter(column) + str(1)].value!=None):
    Column_Names.append(ws1[get_column_letter(column)+str(1)].value)
    column+=1

#This above piece of code is for getting the column names of the input file.

row=2

while (ws1["A" + str(row)].value!=None):
    col=1
    ChapterTest1 = []
    ChapterTest2=[]
    ChapterTest3 = []
    ChapterTest4=[]
    ChapterTest5 = []
    TopicTest1 =[]
    TopicTest2 =[]
    FullChapterTest1= []
    FullChapterTest2= []
    ChapterTestDetails =[]
    TopicTestDetails = []
    FullChapterTestDetails = []
    while (ws1[get_column_letter(col) + str(row)].value!=None):
        if (Column_Names[col-1].__contains__("Concept Test 1")):
            value = ws1[get_column_letter(col) + str(row)].value
            ChapterTest1.append(value)

        if (Column_Names[col-1].__contains__("Concept Test 2")):
            value = ws1[get_column_letter(col) + str(row)].value
            ChapterTest2.append(value)

        if (Column_Names[col-1].__contains__("Concept Test 3")):
            value = ws1[get_column_letter(col) + str(row)].value
            ChapterTest3.append(value)
        
        if (Column_Names[col-1].__contains__("Concept Test 4")):
            value = ws1[get_column_letter(col) + str(row)].value
            ChapterTest4.append(value)
        
        if (Column_Names[col-1].__contains__("Concept Test 5")):
            value = ws1[get_column_letter(col) + str(row)].value
            ChapterTest5.append(value)
        
        if (Column_Names[col-1].__contains__("Topic Test 1")):
            value = ws1[get_column_letter(col) + str(row)].value
            TopicTest1.append(value)

        if (Column_Names[col-1].__contains__("Topic Test 2")):
            value = ws1[get_column_letter(col) + str(row)].value
            TopicTest2.append(value)

        if (Column_Names[col-1].__contains__("Full Chapter Test 1")):
            value = ws1[get_column_letter(col) + str(row)].value
            FullChapterTest1.append(value)

        if (Column_Names[col-1].__contains__("Full Chapter Test 2")):
            value = ws1[get_column_letter(col) + str(row)].value
            FullChapterTest2.append(value)
        
        col+=1
    
    ChapterTestDetails.append(ChapterTest1)
    ChapterTestDetails.append(ChapterTest2)
    ChapterTestDetails.append(ChapterTest3)
    ChapterTestDetails.append(ChapterTest4)
    ChapterTestDetails.append(ChapterTest5)
    ChapterTest.append(ChapterTestDetails)

    TopicTestDetails.append(TopicTest1)
    TopicTestDetails.append(TopicTest2)
    TopicTest.append(TopicTestDetails)

    FullChapterTestDetails.append(FullChapterTest1)
    FullChapterTestDetails.append(FullChapterTest2)
    FullChapterTest.append(FullChapterTestDetails)
    row+=1


#Get_Column_Letter method is used for column number into the column name, it takes number as argument and returns the column name represented in excel.

#__Contins__ method is used in conditional operator for comparing, It takes a string as an argument and returns true if it contains otherwise it returns false.

#The above piece of code is for extracting the data from the input excel file.

for i in range(0,len(ChapterTest)):

    for j in range(0,len(ChapterTest[i])):

        if (ChapterTest[i][j][0]=="-"):
            continue

        appendList= [ws1[get_column_letter(1)+str(i+2)].value,ws1[get_column_letter(2)+str(i+2)].value,ws1[get_column_letter(3)+str(i+2)].value,"Concept Test",ChapterTest[i][j][2],ChapterTest[i][j][3],ChapterTest[i][j][0],ChapterTest[i][j][5],ChapterTest[i][j][1],ChapterTest[i][j][4]]

        ws2.append(appendList) 

    

    for l in range(0,len(FullChapterTest[i])):

        if (FullChapterTest[i][l][0]=="-"):
            continue

        appendList= [ws1[get_column_letter(1)+str(i+2)].value,ws1[get_column_letter(2)+str(i+2)].value,ws1[get_column_letter(3)+str(i+2)].value,"Full Chapter Test",FullChapterTest[i][l][2],FullChapterTest[i][l][3],FullChapterTest[i][l][0],FullChapterTest[i][l][5],FullChapterTest[i][l][1],FullChapterTest[i][l][4]]

        ws2.append(appendList)

    
    for k in range(0,len(TopicTest[i])):

        if (TopicTest[i][k][0]=="-"):
            continue

        appendList= [ws1[get_column_letter(1)+str(i+2)].value,ws1[get_column_letter(2)+str(i+2)].value,ws1[get_column_letter(3)+str(i+2)].value,"Topic Test",TopicTest[i][k][2],TopicTest[i][k][3],TopicTest[i][k][0],TopicTest[i][k][5],TopicTest[i][k][1],TopicTest[i][k][4]]

        ws2.append(appendList)


#This above piece of code is for appending the extracted from input file to the output file.



wb2.save("Output.xlsx")
#Added Comment to practise the git

#Save method is for saving the file created and it takes a string as an argument. The string must the file type and file name that where the output should be saved. 